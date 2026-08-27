"""
Team Trait Comparison pipeline — builds minute-weighted historical G League
team-season profiles (S/W + height/weight/position) for comparison against
the mock roster builder's own aggregate profile.

Data sources (all from Team_Trait_Comp_Database.xlsx):
  - Team Rosters: per-player position/height/weight/team per season
  - Team Minutes: per-player minutes per team per season (the weighting basis)
  - D1 Bart Stats: per-player college stats (feeds S/W, same method site-wide)
  - Team Stats: team-level box-score data (attached as reference context only,
    not part of the similarity computation itself)

Verified team-code mapping: Team Rosters and Team Minutes use DIFFERENT code
schemes for the same teams (e.g. Ignite is "IGN" in Rosters, "GLI" in
Minutes) — confirmed by linking the two schemes via shared players (30/31
links at 100% agreement). Team Stats uses full team names, resolved via a
mix of an official NBA G League source and abbreviation-pattern consistency
checks (all 34 codes verified against the real Team Stats names, zero gaps).
"""

import openpyxl, json, re, datetime
from collections import defaultdict

XLSX_PATH = "Team_Trait_Comp_Database.xlsx"  # place this file in the repo root before running

# ---- Verified code -> team name mapping (Minutes/Team Stats scheme) ----
CODE_TO_TEAM = {
    'AUS': 'Austin Spurs', 'BHM': 'Birmingham Squadron', 'CCG': 'Capital City Go-Go',
    'CLC': 'Cleveland Charge', 'CPS': 'College Park Skyhawks', 'DEL': 'Delaware Blue Coats',
    'GBO': 'Greensboro Swarm', 'GLI': 'G League Ignite', 'GRG': 'Grand Rapids Gold',
    'IMA': 'Indiana Mad Ants', 'IWA': 'Iowa Wolves', 'LIN': 'Long Island Nets',
    'MCC': 'Motor City Cruise', 'MHU': 'Memphis Hustle', 'MNE': 'Maine Celtics',
    'MXC': 'Ciudad de Mexico Capitanes', 'NOB': 'Noblesville Boom', 'OKL': 'Oklahoma City Blue',
    'ONT': 'Ontario Clippers', 'OSC': 'Osceola Magic', 'RAP': 'Raptors 905',
    'RCR': 'Rip City Remix', 'RGV': 'Rio Grande Valley Vipers', 'SBL': 'South Bay Lakers',
    'SCW': 'Santa Cruz Warriors', 'SDC': 'San Diego Clippers', 'SLC': 'Salt Lake City Stars',
    'STO': 'Stockton Kings', 'SXF': 'Sioux Falls Skyforce', 'TEX': 'Texas Legends',
    'VAL': 'Valley Suns', 'WCB': 'Windy City Bulls', 'WES': 'Westchester Knicks',
    'WIS': 'Wisconsin Herd',
}
# Rosters-code -> Minutes-code (verified via shared-player cross-reference, 30/31 at 100%)
ROSTERS_TO_MINUTES_CODE = {
    'AUS': 'AUS', 'BIR': 'BHM', 'CAP': 'MXC', 'CLE': 'CLC', 'COL': 'CPS', 'DEL': 'DEL',
    'GBO': 'GBO', 'GOG': 'CCG', 'GRG': 'GRG', 'IWA': 'IWA', 'LIN': 'LIN', 'MCC': 'MCC',
    'MHU': 'MHU', 'MNE': 'MNE', 'NBL': 'NOB', 'OKL': 'OKL', 'OSC': 'OSC', 'POR': 'RCR',
    'RAP': 'RAP', 'RGV': 'RGV', 'SBL': 'SBL', 'SCW': 'SCW', 'SDC': 'SDC', 'SLC': 'SLC',
    'STO': 'STO', 'SXF': 'SXF', 'TEX': 'TEX', 'VAL': 'VAL', 'WCB': 'WCB', 'WES': 'WES',
    'WIS': 'WIS', 'IGN': 'GLI',
}

SHEET_COLUMNS = [
    {"type": "skip"}, {"type": "text", "key": "class"}, {"type": "text", "key": "height_raw"},
    {"type": "text", "key": "name"}, {"type": "text", "key": "team"}, {"type": "text", "key": "conf"},
    {"type": "num", "key": "gp"}, {"type": "num", "key": "minPct"},
    {"type": "num", "key": "prpg"}, {"type": "num", "key": "dPrpg"},
    {"type": "num", "key": "bpm"}, {"type": "num", "key": "obpm"}, {"type": "num", "key": "dbpm"},
    {"type": "num", "key": "ortg"}, {"type": "num", "key": "drtg"},
    {"type": "num", "key": "usg"}, {"type": "num", "key": "efg"}, {"type": "num", "key": "ts"},
    {"type": "num", "key": "orb"}, {"type": "num", "key": "drb"},
    {"type": "num", "key": "ast"}, {"type": "num", "key": "tov"}, {"type": "num", "key": "astToRatio"},
    {"type": "num", "key": "blk"}, {"type": "num", "key": "stl"},
    {"type": "num", "key": "ftr"}, {"type": "num", "key": "fc40"},
    {"type": "pair", "makes": "dunks", "att": "dunkAttempts", "pct": "dunkPct"},
    {"type": "pair", "makes": "rimMakes", "att": "rimAttempts", "pct": "rimPct"},
    {"type": "pair", "makes": "nonRim2Makes", "att": "nonRim2Attempts", "pct": "nonRim2Pct"},
    {"type": "pair", "makes": "ftMakes", "att": "ftAttempts", "pct": "ftPct"},
    {"type": "pair", "makes": "twoPM", "att": "twoPA", "pct": "twoPct"},
    {"type": "num", "key": "threePRate"}, {"type": "num", "key": "threePer100"},
    {"type": "pair", "makes": "threePM", "att": "threePA", "pct": "threePct"},
]
POS_MAP = {'PG': 'Guard', 'SG': 'Guard', 'G': 'Guard', 'SF': 'Wing', 'GF': 'Wing', 'F': 'Wing',
           'PF': 'Big', 'C': 'Big', 'FC': 'Big', '-': None}
SW_STATS = ["prpg", "dPrpg", "bpm", "obpm", "dbpm", "ortg", "drtg", "usg", "efg", "ts",
            "orb", "drb", "ast", "tov", "astToRatio", "blk", "stl", "ftr",
            "rimPct", "nonRim2Pct", "ftPct", "twoPct", "threePRate", "threePct"]
LOWER_IS_BETTER = {"tov", "drtg"}


SUFFIXES = {"jr", "jr.", "sr", "sr.", "ii", "iii", "iv", "v"}

def normalize(name):
    # Suffix-tolerant: "Kevin Knox II" and "Kevin Knox" must match, since
    # Team Rosters and D1 Bart Stats don't use suffixes consistently for the
    # same real player — confirmed causing real false negatives (Knox,
    # McCullar) in the missing-players export, same fix applied here since
    # this same join feeds the team-season aggregation itself.
    name = re.sub(r"[.,]", "", name.lower())
    words = [w for w in name.split() if w not in SUFFIXES]
    return re.sub(r"\s+", " ", " ".join(words).replace("-", " ")).strip()


def reconstruct_height(raw):
    if isinstance(raw, datetime.datetime):
        feet, inches = raw.month, raw.day
        if 4 <= feet <= 8 and 0 <= inches <= 11:
            return feet * 12 + inches
    return None


def parse_hyphen_pair(cell):
    m = re.match(r"^(\d+)\s*-\s*(\d+)$", str(cell).strip()) if cell is not None else None
    return (int(m.group(1)), int(m.group(2))) if m else (None, None)


def parse_bart_row(cells):
    out = {}; i = 0
    for col in SHEET_COLUMNS:
        if col["type"] == "skip": i += 1; continue
        if col["type"] == "text": out[col["key"]] = cells[i]; i += 1; continue
        if col["type"] == "num":
            v = cells[i]; out[col["key"]] = float(v) if isinstance(v, (int, float)) else None; i += 1; continue
        if col["type"] == "pair":
            makes, att = parse_hyphen_pair(cells[i])
            pct = cells[i+1] if isinstance(cells[i+1], (int, float)) else None
            out[col["makes"]] = makes; out[col["att"]] = att; out[col["pct"]] = pct; i += 2; continue
    return out


def load_workbook_data():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    # Bart Stats lookup, preferring real (non-blank) rows over blank name-only duplicates
    bart_by_name = {}
    for row in wb['D1 Bart Stats'].iter_rows(min_row=2, values_only=True):
        parsed = parse_bart_row(list(row))
        if not parsed.get("name"): continue
        norm = normalize(parsed["name"])
        if norm not in bart_by_name or (parsed["gp"] is not None and bart_by_name[norm]["gp"] is None):
            bart_by_name[norm] = parsed

    # Roster info per (year, normalized name) — single-team rows only (skip mid-season trades)
    roster_info = {}
    for row in wb['Team Rosters'].iter_rows(min_row=2, values_only=True):
        year, player, pos, ht, wt, teams, gp = row[0], row[1], row[2], row[3], row[4], row[5], row[6]
        if not (year and player and teams) or ',' in str(teams):
            continue
        key = (int(year), normalize(player))
        roster_info[key] = {
            'name': player.strip(), 'position': POS_MAP.get(pos),
            'heightIn': reconstruct_height(ht), 'weight': wt,
            'rosters_code': teams.strip(),
        }

    # Minutes per (year, normalized name, minutes-code)
    minutes_by_key = {}
    for row in wb['Team Minutes'].iter_rows(min_row=2, values_only=True):
        year, player, code, mins = row
        if year and player and code and mins is not None:
            minutes_by_key[(int(year), normalize(player))] = (code.strip(), mins)

    # Team Stats: (year, team name) -> full stat row (as a dict via first sub-table's headers)
    ws_stats = wb['Team Stats']
    headers = [c.value for c in ws_stats[1]]
    team_stats_by_key = {}
    for row in ws_stats.iter_rows(min_row=2, values_only=True):
        year, name = row[0], row[1]
        if year and name:
            team_stats_by_key[(int(year), name)] = dict(zip(headers, row))

    return bart_by_name, roster_info, minutes_by_key, team_stats_by_key


def get_sw_value(bart_row, key):
    return bart_row.get(key) if bart_row else None


def build_percentile_pools(roster_info, bart_by_name):
    """One pool per position, of every SW_STATS value across all G League roster
    players (any year) who have real Bart Stats — this is the reference
    population percentiles are computed against."""
    pools = {'Guard': defaultdict(list), 'Wing': defaultdict(list), 'Big': defaultdict(list)}
    for (year, norm_name), info in roster_info.items():
        pos = info['position']
        if not pos or pos not in pools:
            continue
        bart_row = bart_by_name.get(norm_name)
        if not bart_row or bart_row['gp'] is None:
            continue
        for stat_key in SW_STATS:
            val = get_sw_value(bart_row, stat_key)
            if val is not None:
                pools[pos][stat_key].append(val)
    return pools


def percentile_rank(pools, position, stat_key, value):
    if value is None or position not in pools:
        return None
    peers = pools[position].get(stat_key, [])
    if len(peers) < 5:
        return None
    if stat_key in LOWER_IS_BETTER:
        better = sum(1 for v in peers if v >= value)
    else:
        better = sum(1 for v in peers if v <= value)
    return better / len(peers)


def build_team_season_profiles(bart_by_name, roster_info, minutes_by_key, team_stats_by_key, pools):
    """For every (year, team) in Team Stats, aggregate its full roster into a
    minute-weighted S/W profile + height/weight/position mix. Players with no
    usable stats or no minutes entry are excluded from the weighted average
    (not treated as zero) rather than dragging the average toward nothing."""
    # Group roster_info by (year, rosters_code) for lookup
    by_year_code = defaultdict(list)
    for (year, norm_name), info in roster_info.items():
        by_year_code[(year, info['rosters_code'])].append((norm_name, info))

    profiles = []
    years_seen = set(y for (y, _) in team_stats_by_key)

    for year in years_seen:
        for rosters_code, minutes_code in ROSTERS_TO_MINUTES_CODE.items():
            team_name = CODE_TO_TEAM.get(minutes_code)
            if not team_name:
                continue
            stats_key = (year, team_name)
            if stats_key not in team_stats_by_key:
                continue  # this team didn't exist / no data that year

            roster_entries = by_year_code.get((year, rosters_code), [])
            if not roster_entries:
                continue

            weighted_sw = defaultdict(float)
            weighted_sw_weight = defaultdict(float)
            total_min = 0.0
            height_weighted_sum, weight_weighted_sum, min_for_physicals = 0.0, 0.0, 0.0
            position_minutes = defaultdict(float)
            players_included = []
            roster_players = []  # per-player detail, for team-wide queries (e.g. "teams with 2+ players above a stat threshold")

            for norm_name, info in roster_entries:
                min_key = (year, norm_name)
                if min_key not in minutes_by_key:
                    continue
                _, mins = minutes_by_key[min_key]
                total_min += mins
                players_included.append(info['name'])

                if info['position']:
                    position_minutes[info['position']] += mins
                if info['heightIn']:
                    height_weighted_sum += info['heightIn'] * mins
                    min_for_physicals += mins  # shared denominator for height (weight uses its own check below)
                if info['weight'] and isinstance(info['weight'], (int, float)):
                    weight_weighted_sum += info['weight'] * mins

                bart_row = bart_by_name.get(norm_name)
                player_stats = {}
                if bart_row and bart_row['gp'] is not None and info['position']:
                    for stat_key in SW_STATS:
                        raw_value = get_sw_value(bart_row, stat_key)
                        if raw_value is not None:
                            player_stats[stat_key] = raw_value
                        pct = percentile_rank(pools, info['position'], stat_key, raw_value)
                        if pct is not None:
                            weighted_sw[stat_key] += pct * mins
                            weighted_sw_weight[stat_key] += mins

                roster_players.append({
                    'name': info['name'],
                    'position': info['position'],
                    'heightIn': info['heightIn'],
                    'minutes': round(mins),
                    'stats': player_stats,
                })

            if total_min == 0 or not players_included:
                continue

            sw_profile = {}
            for stat_key in SW_STATS:
                if weighted_sw_weight[stat_key] > 0:
                    sw_profile[stat_key] = round(weighted_sw[stat_key] / weighted_sw_weight[stat_key], 4)

            profiles.append({
                'year': year,
                'team': team_name,
                'playersIncluded': len(players_included),
                'totalMinutes': round(total_min),
                'swProfile': sw_profile,
                'avgHeightIn': round(height_weighted_sum / min_for_physicals, 1) if min_for_physicals else None,
                'avgWeight': round(weight_weighted_sum / total_min, 1) if weight_weighted_sum else None,
                'positionMixPct': {pos: round(mins / total_min, 3) for pos, mins in position_minutes.items()},
                'teamStatsReference': team_stats_by_key[stats_key],
                'roster': roster_players,
            })

    return profiles


def build_gleague_reference(roster_info, bart_by_name, minutes_by_key):
    """Regenerates the full G League player coverage reference (used by the
    site's own "Check Missing Stats" feature) from the CURRENT state of the
    workbook — this used to be a separate, one-off script whose output never
    updated after the database changed. Folding it into the same pipeline
    run means it can never drift out of sync with the actual data again."""
    all_gleague = {}
    for (year, norm_name), info in roster_info.items():
        if norm_name not in all_gleague or year > all_gleague[norm_name]['year']:
            all_gleague[norm_name] = {
                'name': info['name'], 'norm': norm_name, 'year': year,
                'team': CODE_TO_TEAM.get(ROSTERS_TO_MINUTES_CODE.get(info['rosters_code'], ''), info['rosters_code']),
            }

    # Total minutes summed across every season on record — the best simple
    # proxy for "how much this player's stats would actually matter to the
    # team-season aggregates," since that's literally what minutes weight
    # in that computation.
    total_minutes_by_norm = {}
    for (year, norm_name), (code, mins) in minutes_by_key.items():
        total_minutes_by_norm[norm_name] = total_minutes_by_norm.get(norm_name, 0) + mins

    output = []
    for norm, info in all_gleague.items():
        bart_row = bart_by_name.get(norm)
        output.append({
            'name': info['name'], 'norm': norm,
            'mostRecentYear': info['year'], 'mostRecentTeam': info['team'],
            'hasBartStats': bool(bart_row and bart_row['gp'] is not None),
            'totalMinutes': round(total_minutes_by_norm.get(norm, 0)),
        })
    output.sort(key=lambda x: -x['totalMinutes'])
    return output


if __name__ == "__main__":
    print("Loading workbook...")
    bart_by_name, roster_info, minutes_by_key, team_stats_by_key = load_workbook_data()
    print(f"Bart Stats (real data): {sum(1 for v in bart_by_name.values() if v['gp'] is not None)}")
    print(f"Roster entries (single-team): {len(roster_info)}")
    print(f"Minutes entries: {len(minutes_by_key)}")
    print(f"Team-season stat rows: {len(team_stats_by_key)}")

    pools = build_percentile_pools(roster_info, bart_by_name)
    for pos in pools:
        sample_stat = SW_STATS[0]
        print(f"{pos} pool for '{sample_stat}': {len(pools[pos][sample_stat])} values")

    print("\nBuilding team-season profiles...")
    profiles = build_team_season_profiles(bart_by_name, roster_info, minutes_by_key, team_stats_by_key, pools)
    print(f"Team-season profiles built: {len(profiles)}")

    years_covered = sorted(set(p['year'] for p in profiles))
    teams_covered = sorted(set(p['team'] for p in profiles))
    print(f"Years covered: {years_covered}")
    print(f"Teams covered: {len(teams_covered)}")

    print("\nRegenerating G League player coverage reference...")
    reference = build_gleague_reference(roster_info, bart_by_name, minutes_by_key)
    with open("data/gleague_players_reference.json", "w") as f:
        json.dump(reference, f, separators=(",", ":"))
    covered = sum(1 for r in reference if r['hasBartStats'])
    print(f"Reference written: {len(reference)} players, {covered} with Bart Stats coverage")

    output_path = "data/team_trait_profiles.json"
    fresh_pools = {pos: {k: v for k, v in stats.items()} for pos, stats in pools.items()}

    # The in-site "Add Players" import writes directly to this same file,
    # appending raw stat values and player names — but this pipeline used to
    # overwrite the whole file from scratch every run, silently destroying
    # anything added that way. Confirmed happening for real: a real user's
    # 171 imported players vanished the moment the Action ran afterward.
    # Both paths only ever grow their pools, never shrink them, so any pool
    # values beyond what this fresh computation produces are preserved as
    # import-contributed and merged back in, rather than discarded.
    imported_names = []
    existing_playtypes_by_key = {}
    try:
        with open(output_path) as f:
            existing = json.load(f)
        imported_names = existing.get('importedPlayerNames', [])
        existing_pools = existing.get('percentilePools', {})
        for pos, stats in existing_pools.items():
            fresh_pools.setdefault(pos, {})
            for stat_key, values in stats.items():
                fresh_count = len(fresh_pools[pos].get(stat_key, []))
                if len(values) > fresh_count:
                    fresh_pools[pos].setdefault(stat_key, [])
                    fresh_pools[pos][stat_key].extend(values[fresh_count:])
        # This pipeline has no source for playtype data at all — it's merged
        # in separately from a different database (playtype leaderboard
        # files), not derivable from Team_Trait_Comp_Database.xlsx. Without
        # explicitly carrying it forward here, every pipeline run would
        # silently wipe it out, the same class of loss already caught once
        # for importedPlayerNames.
        for p in existing.get('profiles', []):
            if p.get('playtypes'):
                existing_playtypes_by_key[(p['year'], p['team'])] = p['playtypes']
        if imported_names:
            print(f"Preserved {len(imported_names)} previously-imported player name(s) from the existing file.")
        if existing_playtypes_by_key:
            print(f"Preserved playtype data for {len(existing_playtypes_by_key)} team-season(s) from the existing file.")
    except (FileNotFoundError, json.JSONDecodeError):
        pass  # no existing file yet, or it's malformed — nothing to preserve, proceed with fresh data only

    fresh_keys = set()
    for p in profiles:
        key = (p['year'], p['team'])
        fresh_keys.add(key)
        if key in existing_playtypes_by_key:
            p['playtypes'] = existing_playtypes_by_key[key]

    # Playtype-only team-seasons (no roster/stats data exists for them, so
    # this pipeline's own computation never produces a profile at all) need
    # to be re-added directly, not just merged into an existing entry.
    for key, playtypes in existing_playtypes_by_key.items():
        if key not in fresh_keys:
            profiles.append({'year': key[0], 'team': key[1], 'playtypes': playtypes})

    output_data = {
        'profiles': profiles,
        'percentilePools': fresh_pools,
        'importedPlayerNames': imported_names,
    }
    with open(output_path, "w") as f:
        json.dump(output_data, f, separators=(",", ":"))
    print(f"\nWritten to {output_path}")
    print(f"Pool sizes exported: { {pos: len(stats.get(SW_STATS[0], [])) for pos, stats in fresh_pools.items()} }")
